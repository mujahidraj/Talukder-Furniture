import openpyxl
import os
import re
import shutil
from datetime import datetime

# ==================== CONFIGURATION ====================

CATALOG_DIR = r"d:\Talukder Furniture\Documents\Product Catalog"
BACKUP_DIR = os.path.join(CATALOG_DIR, "backups")

# Code number → Bengali name mapping
NAME_MAP = {
    # Bedroom Sets
    '109': 'বনলতা',
    '110': 'শুভ্রা',
    '111': 'বিপরীত',
    '112': 'কদম্বরী',
    '113': 'দিগন্ত',
    '114': 'রেখাচিত্র',
    '116': 'অবয়ব',
    '117': 'সূর্যমুখী',
    '118': 'কসমস',
    '119': 'রৌজা',
    '120': 'ফ্লোরাল',
    '201': 'পদ্ম',
    '202': 'শিল্পিক',
    '203': 'অরণ্য',
    '204': 'চারুলতা',
    '205': 'অম্বরী',
    '206': 'পুঞ্জিকা',
    '401': 'মিতালী',
    '402': 'রূপান্তর',
    '403': 'টেরাকোটা',
    '404': 'প্রবালিকা',
    '405': 'মুক্তালী',
    '406': 'জলরেখা',
    '407': 'স্বর্ণালী',
    # Kids Beds
    '131': 'সানরে',
    '132': 'প্রিন্সেস',
    '133': 'হ্যালো কিটি',
    '134': 'স্কাইবো',
    '135': 'পোকো',
    '136': 'মোয়ানা মোয়ানা',
}

# Product type codes that belong to bedroom sets
BEDROOM_TYPE_CODES = {'BED', 'BST', 'DST', 'DTS', 'CBD', 'COD', 'WDR', 'WDD', 'ALN'}

# Files to modify (relative to CATALOG_DIR)
FILES_TO_MODIFY = [
    "10 Bedroom Set.xlsx",
    "10 Bed Room Set-3.xlsx",
    "Bedroom Set Uv Print.xlsx",
    "Kids Bed.xlsx",
    "Dressing Stool.xlsx",
    "4 Product.xlsx",
]

# ==================== HELPER FUNCTIONS ====================

def extract_code_number(product_code):
    """Extract the 3-digit number from a product code like 'TFL-BED-111 LB'"""
    if not product_code:
        return None
    # Match patterns like TFL-BED-111, TFL-BST-204, TFL-DTS-105, etc.
    match = re.search(r'TFL-\w+-(\d{3})', str(product_code))
    if match:
        return match.group(1)
    return None

def extract_type_code(product_code):
    """Extract the type code from a product code like 'TFL-BED-111 LB' → 'BED'"""
    if not product_code:
        return None
    match = re.search(r'TFL-(\w+)-\d{3}', str(product_code))
    if match:
        return match.group(1)
    return None

def standardize_product_name(name):
    """Standardize product names: 'Bed (Double)' → 'Double Bed', etc."""
    if not name:
        return name
    name = str(name).strip()
    
    # Standardize "Bed (Double)" → "Double Bed"
    # Standardize "Bed (Queen)" → "Queen Bed"  
    # Standardize "Bed (King)" → "King Bed"
    match = re.match(r'^Bed\s*\((\w+)\)\s*$', name, re.IGNORECASE)
    if match:
        size = match.group(1).strip()
        return f"{size} Bed"
    
    return name

def build_new_name(bengali_name, product_name):
    """Build the new product name: 'Talukder বিপরীত Double Bed'"""
    standardized = standardize_product_name(product_name)
    return f"Talukder {bengali_name} {standardized}"

# ==================== MAIN LOGIC ====================

def process_file(filepath):
    """Process a single Excel file and rename products."""
    changes = []
    
    wb = openpyxl.load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Determine the code column and name column
        # Based on our analysis: Code is in Col B, Name is in Col C (for most files)
        # But some files have Code in Col B (index 2) and Name in Col C (index 3)
        # Let's detect from headers
        code_col = None
        name_col = None
        
        # Check first 5 rows for header
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
            for cell in row:
                if cell.value and 'Product Code' in str(cell.value):
                    code_col = cell.column
                if cell.value and 'Product Name' in str(cell.value):
                    name_col = cell.column
        
        if code_col is None or name_col is None:
            print(f"  WARNING: Could not find code/name columns in sheet '{sheet_name}' of {os.path.basename(filepath)}")
            continue
        
        print(f"  Sheet: '{sheet_name}' | Code col: {code_col} | Name col: {name_col}")
        
        # Process each row
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            code_cell = None
            name_cell = None
            for cell in row:
                if cell.column == code_col:
                    code_cell = cell
                if cell.column == name_col:
                    name_cell = cell
            
            if not code_cell or not code_cell.value or not name_cell or not name_cell.value:
                continue
            
            product_code = str(code_cell.value).strip()
            product_name = str(name_cell.value).strip()
            
            # Extract the type code and number
            type_code = extract_type_code(product_code)
            code_number = extract_code_number(product_code)
            
            if not type_code or not code_number:
                continue
            
            # Check if this is a bedroom product type
            if type_code not in BEDROOM_TYPE_CODES:
                continue
            
            # Check if we have a name mapping for this code
            if code_number not in NAME_MAP:
                continue
            
            bengali_name = NAME_MAP[code_number]
            new_name = build_new_name(bengali_name, product_name)
            
            # Check if already renamed (avoid double-renaming)
            if product_name.startswith('Talukder '):
                continue
            
            changes.append({
                'sheet': sheet_name,
                'row': code_cell.row,
                'code': product_code,
                'old_name': product_name,
                'new_name': new_name,
            })
            
            # Apply the change
            name_cell.value = new_name
    
    if changes:
        wb.save(filepath)
        print(f"  ✅ Saved {len(changes)} changes")
    else:
        print(f"  ℹ️  No changes needed")
    
    wb.close()
    return changes

# ==================== EXECUTION ====================

print("=" * 80)
print("BEDROOM PRODUCT RENAMING SCRIPT")
print("=" * 80)

# Step 1: Create backups
print("\n📁 Creating backups...")
os.makedirs(BACKUP_DIR, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

for filename in FILES_TO_MODIFY:
    src = os.path.join(CATALOG_DIR, filename)
    if os.path.exists(src):
        backup_name = f"{os.path.splitext(filename)[0]}_backup_{timestamp}{os.path.splitext(filename)[1]}"
        dst = os.path.join(BACKUP_DIR, backup_name)
        shutil.copy2(src, dst)
        print(f"  ✅ Backed up: {filename}")
    else:
        print(f"  ⚠️  Not found: {filename}")

# Step 2: Process files
print("\n📝 Processing files...")
all_changes = {}

for filename in FILES_TO_MODIFY:
    filepath = os.path.join(CATALOG_DIR, filename)
    if not os.path.exists(filepath):
        print(f"\n⚠️  Skipping (not found): {filename}")
        continue
    
    print(f"\n{'─' * 60}")
    print(f"📄 {filename}")
    print(f"{'─' * 60}")
    
    changes = process_file(filepath)
    all_changes[filename] = changes

# Step 3: Print summary report
print("\n\n" + "=" * 80)
print("📊 SUMMARY REPORT")
print("=" * 80)

total_changes = 0
for filename, changes in all_changes.items():
    print(f"\n📄 {filename}: {len(changes)} products renamed")
    if changes:
        for c in changes:
            print(f"   Row {c['row']:3d} | {c['code']:25s} | {c['old_name']:25s} → {c['new_name']}")
    total_changes += len(changes)

print(f"\n{'=' * 80}")
print(f"✅ TOTAL: {total_changes} products renamed across {len(all_changes)} files")
print(f"📁 Backups saved to: {BACKUP_DIR}")
print(f"{'=' * 80}")
